import io
import re
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from app.models import LeadCreate

COLUMN_ALIASES = {
    "name": {"name", "имя", "фио", "клиент", "client", "заявка"},
    "email": {"email", "e_mail", "mail", "почта"},
    "phone": {"phone", "телефон", "тел", "mobile"},
    "contact": {"contact", "контакт"},
    "source": {"source", "источник", "канал", "откуда"},
    "notes": {"notes", "note", "comment", "комментарий", "описание", "примечание"},
}

COLUMN_KEYWORDS = {
    "name": ("имя", "name", "фио", "клиент", "client"),
    "email": ("email", "e_mail", "почт", "mail"),
    "phone": ("телефон", "phone", "тел", "mobile"),
    "contact": ("контакт", "contact"),
    "source": ("источник", "source", "канал"),
    "notes": ("комментарий", "comment", "note", "описан", "примеч"),
}

BITRIX_MARKERS = {"название_лида", "рабочий_телефон", "источник"}
BITRIX_PHONE_COLUMNS = (
    "Рабочий телефон",
    "Мобильный телефон",
    "Другой телефон",
    "Домашний телефон",
    "Номер факса",
)
BITRIX_EMAIL_COLUMNS = (
    "Рабочий e-mail",
    "Частный e-mail",
    "E-mail для рассылок",
    "Другой e-mail",
)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\ufeff", "")
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"[^a-zа-яё0-9_]", "", text)
    return text


def _cell_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _header_index(headers: list[Any], title: str) -> Optional[int]:
    target = _normalize_header(title)
    for index, header in enumerate(headers):
        if _normalize_header(header) == target:
            return index
    return None


def _is_bitrix_export(headers: list[Any]) -> bool:
    normalized = {_normalize_header(header) for header in headers if header is not None}
    return BITRIX_MARKERS.issubset(normalized)


def _match_header_to_field(normalized: str) -> Optional[str]:
    if not normalized:
        return None

    for field, aliases in COLUMN_ALIASES.items():
        if normalized in aliases:
            return field

    for field, keywords in COLUMN_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            return field

    return None


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        field = _match_header_to_field(_normalize_header(header))
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    best_row_index = 0
    best_mapping: dict[str, int] = {}
    best_score = 0

    for row_index, row in enumerate(rows[:20]):
        mapping = _map_headers(list(row))
        score = len(mapping)
        if "name" in mapping:
            score += 2
        if score > best_score:
            best_score = score
            best_row_index = row_index
            best_mapping = mapping

    return best_row_index, best_mapping


def _headers_preview(headers: list[Any]) -> str:
    cells = [_cell_value(cell) for cell in headers if _cell_value(cell)]
    return ", ".join(cells[:10]) if cells else "пусто"


def _resolve_email(
    email: Optional[str],
    contact: Optional[str],
) -> Optional[str]:
    if email:
        return email
    if contact and "@" in contact:
        return contact
    return None


def _resolve_phone(phone: Optional[str], contact: Optional[str]) -> Optional[str]:
    if phone:
        return phone
    if contact and "@" not in contact:
        return contact
    return None


def _first_value(row: tuple[Any, ...], headers: list[Any], titles: tuple[str, ...]) -> Optional[str]:
    for title in titles:
        index = _header_index(headers, title)
        if index is None or index >= len(row):
            continue
        value = _cell_value(row[index])
        if value:
            return value
    return None


def _bitrix_name(row: tuple[Any, ...], headers: list[Any]) -> Optional[str]:
    parts = [
        _first_value(row, headers, ("Имя",)),
        _first_value(row, headers, ("Отчество",)),
        _first_value(row, headers, ("Фамилия",)),
    ]
    full_name = " ".join(part for part in parts if part)
    if full_name:
        return full_name
    return _first_value(row, headers, ("Название лида",))


def _bitrix_notes(row: tuple[Any, ...], headers: list[Any]) -> Optional[str]:
    parts = [
        _first_value(row, headers, ("Комментарий",)),
        _first_value(row, headers, ("Обращение",)),
        _first_value(row, headers, ("Стадия",)),
        _first_value(row, headers, ("Дополнительно о стадии",)),
    ]
    notes = " | ".join(part for part in parts if part)
    return notes or None


def _parse_bitrix_rows(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
) -> tuple[list[LeadCreate], list[str]]:
    headers = list(rows[header_row_index])
    leads: list[LeadCreate] = []
    errors: list[str] = []

    for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue

        name = _bitrix_name(row, headers)
        if not name:
            errors.append(f"Строка {row_index}: не указано имя")
            continue

        phone = _first_value(row, headers, BITRIX_PHONE_COLUMNS)
        email_raw = _first_value(row, headers, BITRIX_EMAIL_COLUMNS)
        source = _first_value(row, headers, ("Источник",))
        notes = _bitrix_notes(row, headers)

        try:
            email = _resolve_email(email_raw, contact=None)
            lead = LeadCreate(
                name=name,
                contact=phone or email_raw or "",
                comment=notes or "",
                email=email,
                phone=phone,
                source=source,
                notes=notes,
            )
            leads.append(lead)
        except ValidationError as exc:
            details = "; ".join(
                f"{err['loc'][0]}: {err['msg']}" for err in exc.errors()
            )
            errors.append(f"Строка {row_index}: {details}")

    return leads, errors


def _pick_data_sheet(workbook) -> tuple[Any, list[tuple[Any, ...]]]:
    best_sheet = workbook.active
    best_rows: list[tuple[Any, ...]] = []
    best_score = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = list(rows[0])
        if not _is_bitrix_export(headers):
            continue

        score = len(rows)
        if score > best_score:
            best_score = score
            best_sheet = sheet
            best_rows = rows

    if best_rows:
        return best_sheet, best_rows

    sheet = workbook.active
    return sheet, list(sheet.iter_rows(values_only=True))


def parse_xlsx_rows(file_bytes: bytes) -> tuple[list[LeadCreate], list[str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    _, rows = _pick_data_sheet(workbook)
    workbook.close()

    if not rows:
        raise ValueError("Файл пустой")

    header_row_index, column_map = _find_header_row(rows)
    headers = list(rows[header_row_index])

    if _is_bitrix_export(headers):
        leads, errors = _parse_bitrix_rows(rows, header_row_index)
    else:
        if "name" not in column_map:
            preview = _headers_preview(headers)
            raise ValueError(
                "Не найдена колонка с именем. "
                f"Первая строка файла: «{preview}». "
                "Нужна колонка: Имя, ФИО, Клиент или name"
            )

        leads = []
        errors = []
        data_rows = rows[header_row_index + 1 :]

        if not data_rows:
            raise ValueError("В файле есть только заголовки, нет строк с данными")

        for row_index, row in enumerate(data_rows, start=header_row_index + 2):
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue

            def get_field(field: str) -> Optional[str]:
                if field not in column_map:
                    return None
                idx = column_map[field]
                if idx >= len(row):
                    return None
                return _cell_value(row[idx])

            name = get_field("name")
            if not name:
                errors.append(f"Строка {row_index}: не указано имя")
                continue

            email_raw = get_field("email")
            contact = get_field("contact")
            phone_raw = get_field("phone")

            try:
                email = _resolve_email(email_raw, contact)
                phone = _resolve_phone(phone_raw, contact)
                resolved_contact = contact or phone or email_raw or ""
                notes = get_field("notes")
                lead = LeadCreate(
                    name=name,
                    contact=resolved_contact,
                    comment=notes or "",
                    email=email,
                    phone=phone,
                    source=get_field("source"),
                    notes=notes,
                )
                leads.append(lead)
            except ValidationError as exc:
                details = "; ".join(
                    f"{err['loc'][0]}: {err['msg']}" for err in exc.errors()
                )
                errors.append(f"Строка {row_index}: {details}")

    if not leads and errors:
        preview = "; ".join(errors[:5])
        if len(errors) > 5:
            preview += f" ... и ещё {len(errors) - 5}"
        raise ValueError(f"Не удалось разобрать строки: {preview}")

    if not leads:
        raise ValueError(
            "Не найдено строк с данными. Проверьте, что под заголовками есть заполненные строки"
        )

    return leads, errors


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(["Имя", "Email", "Телефон", "Источник", "Комментарий"])
    sheet.append(
        [
            "Ирина Петрова",
            "irina@example.com",
            "+79990000001",
            "landing",
            "Хочу консультацию",
        ]
    )
    sheet.append(
        [
            "Алексей Смирнов",
            "",
            "+79991112233",
            "telegram",
            "Только телефон — email можно оставить пустым",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
